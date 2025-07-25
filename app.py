from flask import Flask, render_template, request, redirect, url_for, session
import openpyxl

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Secret key for session management

# Load or create Excel sheets
try:
    users_wb = openpyxl.load_workbook("users2.xlsx")
except FileNotFoundError:
    users_wb = openpyxl.Workbook()
    users_ws = users_wb.active
    users_ws.append(["Username", "Password", "Email", "Phone", "SkillsOffered", "SkillsWanted"])
    users_wb.save("users2.xlsx")

try:
    skills_wb = openpyxl.load_workbook("skills.xlsx")
except FileNotFoundError:
    skills_wb = openpyxl.Workbook()
    skills_ws = skills_wb.active
    skills_ws.append(["Skill", "Username"])
    skills_wb.save("skills.xlsx")

try:
    sessions_wb = openpyxl.load_workbook("sessions.xlsx")
except FileNotFoundError:
    sessions_wb = openpyxl.Workbook()
    sessions_ws = sessions_wb.active
    sessions_ws.append(["User", "Teacher", "Date", "Time"])
    sessions_wb.save("sessions.xlsx")

users_ws = users_wb.active
skills_ws = skills_wb.active
sessions_ws = sessions_wb.active


@app.route('/')
def home():
    return render_template('index.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        email = request.form['email']
        phone = request.form['phone']
        skill_offered = request.form['skill_offered']
        skill_needed = request.form['skill_needed']

        # Check if the username already exists
        for row in users_ws.iter_rows(values_only=True):
            if row[0] == username:  # Check if username already exists
                return render_template('signup.html', error="Username already exists. Please choose a different one.")

        # Save to users.xlsx with email and phone
        users_ws.append([username, password, email, phone, skill_offered, skill_needed])
        users_wb.save("users2.xlsx")

        # Save skills to skills.xlsx
        skills_ws.append([skill_offered, username])
        skills_ws.append([skill_needed, username])
        skills_wb.save("skills.xlsx")

        return redirect(url_for('login'))
    return render_template('signup.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # Validate credentials
        for row in users_ws.iter_rows(values_only=True):
            if row[0] == username and row[1] == password:
                session['username'] = username  # Store user in session
                return redirect(url_for('dashboard', user=username))
        return "Invalid credentials. Try again."
    return render_template('login.html')


@app.route('/dashboard/<user>', methods=['GET', 'POST'])
def dashboard(user):
    user_data = None
    matches = []

    # Fetch user data
    for row in users_ws.iter_rows(values_only=True):
        if row[0] == user:
            user_data = row
            break

    # If no user data was found, return an error
    if user_data is None:
        return "User not found", 404

    # Matching logic: check if the skills match in reverse
    for row in users_ws.iter_rows(values_only=True):
        if row[0] != user:  # Exclude the current user
            if (row[4] == user_data[5] and row[5] == user_data[4]):  # Match the skills
                matches.append({"username": row[0], "skill_offered": row[4], "skill_needed": row[5]})

    # Check if a success message exists
    success_message = request.args.get('success_message')

    return render_template('dashboard.html', user=user, matches=matches, success_message=success_message)


@app.route('/match', methods=['GET', 'POST'])
def match():
    user = session.get('username')  # Get the logged-in user's username from session
    if request.method == 'POST':
        offered_skills = request.form['offered_skills'].strip().lower()  # Normalize input
        needed_skills = request.form['needed_skills'].strip().lower()  # Normalize input
        matches = []

        # Convert skills to sets for easier comparison (handles multiple skills)
        offered_skills_set = set(offered_skills.split(','))
        needed_skills_set = set(needed_skills.split(','))

        # Debugging: Check user-entered skills
        print(f"Offered Skills Set: {offered_skills_set}")
        print(f"Needed Skills Set: {needed_skills_set}")

        # Search for matches in the database
        for row in users_ws.iter_rows(values_only=True):
            user_offered_skills_set = set(row[4].strip().lower().split(','))
            user_needed_skills_set = set(row[5].strip().lower().split(','))

            # Debugging: Print each user's skills
            print(f"User: {row[0]} | Offered: {user_offered_skills_set}, Needed: {user_needed_skills_set}")

            # Check if skills match
            if needed_skills_set & user_offered_skills_set and offered_skills_set & user_needed_skills_set:
                matches.append({"username": row[0], "skill_offered": row[4], "skill_needed": row[5]})

        # Debugging: Print final matches
        print(f"Matches Found: {matches}")

        if matches:
            return render_template('match.html', matches=matches, user=user)  # Pass user to match.html
        else:
            return render_template('match.html', matches=[], message="No matches found.", user=user)  # Pass user to match.html

    return render_template('match.html', matches=[], message="Please enter your skills.", user=user)  # Pass user to match.html


@app.route('/book_session/<teacher>/<user>', methods=['GET', 'POST'])
def book_session(teacher, user):
    if request.method == 'POST':
        # Retrieve session details from the form
        session_date = request.form['session_date']
        session_time = request.form['session_time']

        # Save session details in the `sessions.xlsx` file
        sessions_ws.append([user, teacher, session_date, session_time])
        sessions_wb.save("sessions.xlsx")

        # Redirect to the mode selection page
        return redirect(url_for('mode', user=user, teacher=teacher))

    return render_template('book_session.html', teacher=teacher, user=user)

@app.route('/mode/<user>/<teacher>', methods=['GET', 'POST'])
def mode(user, teacher):
    if request.method == 'POST':
        mode_of_class = request.form.get('mode')  # Get the selected mode

        # Save the mode or handle it as necessary
        print(f"Mode selected by {user}: {mode_of_class}")

        # Render a confirmation page with the message
        return render_template(
            'confirmation.html',
            user=user,
            teacher=teacher,
            mode=mode_of_class,
            message="Your booking has been confirmed!"
        )

    return render_template('mode.html', user=user, teacher=teacher)



if __name__ == "__main__":
    app.run(debug=True)
